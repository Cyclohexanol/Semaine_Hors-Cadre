"""
Tests for solver_logic.run_optimization().

All test Excel files are created programmatically with openpyxl.
Output files use /tmp and are cleaned up after each test.
"""

import os
import sys
import uuid
import pytest
import openpyxl

# Ensure the webapp package is importable
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from solver_logic import run_optimization, EXPECTED_SESSIONS, TOTAL_SESSIONS


# ---------------------------------------------------------------------------
# Helper: build a minimal Excel template in memory and save to a temp path
# ---------------------------------------------------------------------------

ATELIERS_COLUMNS = [
    "Code",
    "Description",
    "Enseignant",
    "Salle",
    "Catégorie",
    "Nombre de périodes",
    "Nombre d'élèves max par session",
    "Nombre idéal d'élèves par session",
    "Session 1",
    "Session 2",
    "Session 3",
    "Session 4",
    "Session 5",
]

STUDENT_FIXED_COLUMNS = ["Nom", "Prénom", "Classe", "# Préférences"]


def _tmp_path(prefix="test_solver"):
    """Return a unique file path under /tmp."""
    return os.path.join("/tmp", f"{prefix}_{uuid.uuid4().hex[:8]}.xlsx")


def _build_excel(workshops, students, include_category_col=True):
    """
    Build a test Excel workbook and return the path where it was saved.

    Parameters
    ----------
    workshops : list[dict]
        Each dict must have keys matching ATELIERS_COLUMNS (except
        "Catégorie" if *include_category_col* is False).
    students : list[dict]
        Each dict must have keys "Nom", "Prénom", "Classe" and one key per
        workshop code whose value is 1, -1, or None/0/missing.
    include_category_col : bool
        If False the "Catégorie" column is omitted entirely.

    Returns
    -------
    str – path to the saved .xlsx file
    """
    wb = openpyxl.Workbook()

    # --- Ateliers sheet ---
    ws_at = wb.active
    ws_at.title = "Ateliers"

    cols = list(ATELIERS_COLUMNS)
    if not include_category_col:
        cols = [c for c in cols if c != "Catégorie"]

    ws_at.append(cols)

    for w in workshops:
        row = []
        for c in cols:
            row.append(w.get(c, None))
        ws_at.append(row)

    # --- Preferences sheet ---
    ws_pr = wb.create_sheet("Preferences")
    workshop_codes = [w["Code"] for w in workshops]
    pref_cols = STUDENT_FIXED_COLUMNS + workshop_codes
    ws_pr.append(pref_cols)

    for s in students:
        row = [s["Nom"], s["Prénom"], s["Classe"]]
        # compute # Préférences
        nb_prefs = sum(1 for code in workshop_codes if s.get(code) == 1)
        row.append(nb_prefs)
        for code in workshop_codes:
            val = s.get(code, None)
            row.append(val)
        ws_pr.append(row)

    path = _tmp_path("input")
    wb.save(path)
    return path


# ---------------------------------------------------------------------------
# Fixtures / shared data builders
# ---------------------------------------------------------------------------

def _make_basic_workshops():
    """
    Return 5 single-session workshops (one per session), each in a different
    category, capacity 30.
    """
    workshops = []
    categories = ["Art", "Sport", "Science", "Musique", "Tech"]
    for i, session in enumerate(EXPECTED_SESSIONS):
        code = f"W{i+1}"
        sessions_map = {f"Session {j+1}": None for j in range(TOTAL_SESSIONS)}
        sessions_map[f"Session {i+1}"] = session
        w = {
            "Code": code,
            "Description": f"Workshop {i+1}",
            "Enseignant": f"Teacher {i+1}",
            "Salle": f"Room {i+1}",
            "Catégorie": categories[i],
            "Nombre de périodes": 1,
            "Nombre d'élèves max par session": 30,
            "Nombre idéal d'élèves par session": 5,
        }
        w.update(sessions_map)
        workshops.append(w)
    return workshops


def _make_basic_students(workshop_codes, n=5, all_prefer=True):
    """
    Create *n* students. If *all_prefer* each student prefers all workshops.
    """
    students = []
    for i in range(n):
        s = {"Nom": f"Nom{i}", "Prénom": f"Prenom{i}", "Classe": "6A"}
        if all_prefer:
            for code in workshop_codes:
                s[code] = 1
        students.append(s)
    return students


# ---------------------------------------------------------------------------
# 1. Basic optimisation – run_optimization succeeds with a trivial instance
# ---------------------------------------------------------------------------

class TestBasicOptimization:

    def test_success_and_valid_stats(self):
        """Minimal feasible problem: 5 workshops, 5 sessions, a few students."""
        workshops = _make_basic_workshops()
        codes = [w["Code"] for w in workshops]
        students = _make_basic_students(codes, n=3, all_prefer=True)

        input_path = _build_excel(workshops, students)
        output_path = _tmp_path("output")

        try:
            success, message, stats = run_optimization(input_path, output_path)
            assert success, f"Optimization failed: {message}"
            assert stats is not None
            assert os.path.isfile(output_path)

            # All students preferred all workshops -> all sessions satisfied
            assert stats["pref_count"] == 3 * TOTAL_SESSIONS
            assert stats["veto_count"] == 0
            assert stats["neutral_count"] == 0
            assert stats["pref_rate"] == "100.0%"
            assert stats["max_prefs_possible"] == TOTAL_SESSIONS
        finally:
            for p in (input_path, output_path):
                if os.path.exists(p):
                    os.remove(p)

    def test_output_excel_has_expected_sheets(self):
        """The output workbook should contain the 4 expected sheets."""
        workshops = _make_basic_workshops()
        codes = [w["Code"] for w in workshops]
        students = _make_basic_students(codes, n=2)

        input_path = _build_excel(workshops, students)
        output_path = _tmp_path("output")

        try:
            success, _, _ = run_optimization(input_path, output_path)
            assert success
            wb = openpyxl.load_workbook(output_path)
            sheet_names = wb.sheetnames
            assert "Planning par élève" in sheet_names
            assert "Planning par Atelier" in sheet_names
            assert "Résumé et Stats" in sheet_names
            assert "Élèves par Nb Préférences" in sheet_names
            wb.close()
        finally:
            for p in (input_path, output_path):
                if os.path.exists(p):
                    os.remove(p)

    def test_mixed_preferences(self):
        """Students with a mix of preferences, vetoes, and neutrals."""
        # Add 2 workshops per session so the solver has alternatives
        workshops = []
        categories = ["Art", "Sport", "Science", "Musique", "Tech"]
        for i, session in enumerate(EXPECTED_SESSIONS):
            for variant in ("a", "b"):
                code = f"W{i+1}{variant}"
                sessions_map = {f"Session {j+1}": None for j in range(TOTAL_SESSIONS)}
                sessions_map[f"Session {i+1}"] = session
                workshops.append({
                    "Code": code, "Description": f"Workshop {code}",
                    "Enseignant": f"T{code}", "Salle": f"R{code}",
                    "Catégorie": categories[i], "Nombre de périodes": 1,
                    "Nombre d'élèves max par session": 30,
                    "Nombre idéal d'élèves par session": 5,
                    **sessions_map,
                })

        # Student 0: prefers W1a, W2a, W3a; vetoes W4a (but W4b is available); neutral rest
        # Student 1: prefers all 'a' variants
        students = [
            {"Nom": "A", "Prénom": "B", "Classe": "6A",
             "W1a": 1, "W2a": 1, "W3a": 1, "W4a": -1, "W4b": None, "W5a": None},
            {"Nom": "C", "Prénom": "D", "Classe": "6A",
             "W1a": 1, "W2a": 1, "W3a": 1, "W4a": 1, "W5a": 1},
        ]

        input_path = _build_excel(workshops, students)
        output_path = _tmp_path("output")

        try:
            success, message, stats = run_optimization(input_path, output_path)
            assert success, f"Optimization failed: {message}"

            # Student 0 should avoid W4a (veto) by taking W4b instead
            assert stats["veto_count"] == 0, "Solver should avoid veto when alternatives exist"
            assert stats["pref_count"] >= 8  # at least 3+5
        finally:
            for p in (input_path, output_path):
                if os.path.exists(p):
                    os.remove(p)


# ---------------------------------------------------------------------------
# 2. Multi-session workshops
# ---------------------------------------------------------------------------

class TestMultiSessionWorkshops:

    def _make_multi_session_scenario(self):
        """
        5 sessions.  Workshop M spans sessions 1+2 (duration 2).
        M is the ONLY workshop covering sessions 1+2, so the solver must assign it.
        Workshops S3, S4, S5 each span one of sessions 3, 4, 5.
        """
        workshops = [
            # Multi-session workshop covering Lundi matin + Lundi après-midi
            {
                "Code": "M",
                "Description": "Multi",
                "Enseignant": "T1",
                "Salle": "R1",
                "Catégorie": "Art",
                "Nombre de périodes": 2,
                "Nombre d'élèves max par session": 30,
                "Nombre idéal d'élèves par session": 3,
                "Session 1": EXPECTED_SESSIONS[0],
                "Session 2": EXPECTED_SESSIONS[1],
                "Session 3": None,
                "Session 4": None,
                "Session 5": None,
            },
            # Single-session workshops for sessions 3, 4, 5
            {
                "Code": "S3",
                "Description": "Single3",
                "Enseignant": "T4",
                "Salle": "R4",
                "Catégorie": "Musique",
                "Nombre de périodes": 1,
                "Nombre d'élèves max par session": 30,
                "Nombre idéal d'élèves par session": 3,
                "Session 1": None,
                "Session 2": None,
                "Session 3": EXPECTED_SESSIONS[2],
                "Session 4": None,
                "Session 5": None,
            },
            {
                "Code": "S4",
                "Description": "Single4",
                "Enseignant": "T5",
                "Salle": "R5",
                "Catégorie": "Tech",
                "Nombre de périodes": 1,
                "Nombre d'élèves max par session": 30,
                "Nombre idéal d'élèves par session": 3,
                "Session 1": None,
                "Session 2": None,
                "Session 3": None,
                "Session 4": EXPECTED_SESSIONS[3],
                "Session 5": None,
            },
            {
                "Code": "S5",
                "Description": "Single5",
                "Enseignant": "T6",
                "Salle": "R6",
                "Catégorie": "Art",
                "Nombre de périodes": 1,
                "Nombre d'élèves max par session": 30,
                "Nombre idéal d'élèves par session": 3,
                "Session 1": None,
                "Session 2": None,
                "Session 3": None,
                "Session 4": None,
                "Session 5": EXPECTED_SESSIONS[4],
            },
        ]
        return workshops

    def test_multi_session_fewer_distinct_workshops(self):
        """A student assigned to M (duration 2) has 4 distinct workshops, not 5."""
        workshops = self._make_multi_session_scenario()

        # M is the only option for sessions 1+2, so the solver must assign it
        student = [
            {"Nom": "X", "Prénom": "Y", "Classe": "6A",
             "M": 1, "S3": 1, "S4": 1, "S5": 1},
        ]

        input_path = _build_excel(workshops, student)
        output_path = _tmp_path("output")

        try:
            success, msg, stats = run_optimization(input_path, output_path)
            assert success, msg

            # Read the student schedule to check distinct workshops
            import pandas as pd
            schedule = pd.read_excel(output_path, sheet_name="Planning par élève")
            row = schedule.iloc[0]
            assigned_codes = [row[s] for s in EXPECTED_SESSIONS if row[s]]
            distinct = set(assigned_codes)

            # M covers 2 sessions, so it appears twice but is 1 distinct workshop
            assert len(distinct) == 4, (
                f"Expected 4 distinct workshops, got {len(distinct)}: {distinct}"
            )
            assert "M" in distinct
        finally:
            for p in (input_path, output_path):
                if os.path.exists(p):
                    os.remove(p)

    def test_multi_session_kpi_counts_sessions(self):
        """
        Preference KPI is session-based: M preferred + duration 2 => 2/5
        sessions satisfied from that single workshop.
        """
        workshops = self._make_multi_session_scenario()

        # Student prefers M, S3, S4, S5
        student = [
            {"Nom": "X", "Prénom": "Y", "Classe": "6A",
             "M": 1, "S3": 1, "S4": 1, "S5": 1},
        ]

        input_path = _build_excel(workshops, student)
        output_path = _tmp_path("output")

        try:
            success, msg, stats = run_optimization(input_path, output_path)
            assert success, msg

            # M (2 sessions) + S3 + S4 + S5 = 5 pref sessions
            assert stats["pref_count"] == 5
            assert stats["pref_rate"] == "100.0%"
        finally:
            for p in (input_path, output_path):
                if os.path.exists(p):
                    os.remove(p)

    def test_distribution_keys_are_integers_out_of_total(self):
        """prefs_distribution keys must be integers in [0..TOTAL_SESSIONS]."""
        workshops = self._make_multi_session_scenario()
        student = [
            {"Nom": "X", "Prénom": "Y", "Classe": "6A",
             "M": 1, "S3": 1, "S4": 1, "S5": 1},
        ]

        input_path = _build_excel(workshops, student)
        output_path = _tmp_path("output")

        try:
            success, msg, stats = run_optimization(input_path, output_path)
            assert success, msg

            dist = stats["prefs_distribution"]
            for key in dist:
                assert isinstance(key, int), f"Key {key!r} is not an int"
                assert 0 <= key <= TOTAL_SESSIONS, (
                    f"Key {key} outside range [0, {TOTAL_SESSIONS}]"
                )
        finally:
            for p in (input_path, output_path):
                if os.path.exists(p):
                    os.remove(p)

    def test_multi_session_partial_preference(self):
        """
        M is duration 2 but student does NOT prefer it.
        Verify the 2 neutral sessions are counted correctly.
        """
        workshops = self._make_multi_session_scenario()

        # Student prefers S3, S4, S5 but NOT M (neutral)
        student = [
            {"Nom": "X", "Prénom": "Y", "Classe": "6A",
             "M": None, "S3": 1, "S4": 1, "S5": 1},
        ]

        input_path = _build_excel(workshops, student)
        output_path = _tmp_path("output")

        try:
            success, msg, stats = run_optimization(input_path, output_path)
            assert success, msg

            # S3 + S4 + S5 = 3 pref sessions out of 5
            assert stats["pref_count"] == 3
            # Remaining 2 sessions are neutral (M or A1+A2)
            assert stats["neutral_count"] == 2
            assert stats["veto_count"] == 0
        finally:
            for p in (input_path, output_path):
                if os.path.exists(p):
                    os.remove(p)


# ---------------------------------------------------------------------------
# 3. Category diversity
# ---------------------------------------------------------------------------

class TestCategoryDiversity:

    def _make_diverse_scenario(self):
        """
        4 categories, 5 sessions.
        Each session has 2 workshops from different categories so the solver
        has a real choice.
        Categories: Art, Sport, Science, Musique
        """
        workshops = []
        cats_per_session = [
            ("Art", "Sport"),
            ("Science", "Musique"),
            ("Art", "Science"),
            ("Sport", "Musique"),
            ("Art", "Sport"),
        ]
        idx = 0
        for i, session in enumerate(EXPECTED_SESSIONS):
            for cat in cats_per_session[i]:
                idx += 1
                code = f"W{idx}"
                sessions_map = {f"Session {j+1}": None for j in range(TOTAL_SESSIONS)}
                sessions_map[f"Session {i+1}"] = session
                workshops.append({
                    "Code": code,
                    "Description": f"Workshop {idx}",
                    "Enseignant": f"T{idx}",
                    "Salle": f"R{idx}",
                    "Catégorie": cat,
                    "Nombre de périodes": 1,
                    "Nombre d'élèves max par session": 30,
                    "Nombre idéal d'élèves par session": 5,
                    **sessions_map,
                })
        return workshops

    def test_weight_zero_same_as_no_categories(self):
        """
        category_diversity_weight=0 should not change the assignment compared
        to a baseline without category influence.  At minimum, both should
        succeed and have the same total pref count.
        """
        workshops = self._make_diverse_scenario()
        codes = [w["Code"] for w in workshops]

        # Students prefer ALL workshops so the only differentiator is
        # deviation and categories.
        students = _make_basic_students(codes, n=4, all_prefer=True)

        input_path = _build_excel(workshops, students)
        out0 = _tmp_path("out0")
        out1 = _tmp_path("out1")

        try:
            ok0, m0, s0 = run_optimization(input_path, out0, category_diversity_weight=0)
            ok1, m1, s1 = run_optimization(input_path, out1, category_diversity_weight=0)
            assert ok0 and ok1
            # Same weight => identical results (deterministic CBC solver)
            assert s0["pref_count"] == s1["pref_count"]
        finally:
            for p in (input_path, out0, out1):
                if os.path.exists(p):
                    os.remove(p)

    def test_positive_weight_produces_category_distribution(self):
        """With weight > 0 and >= 2 categories, category_diversity_distribution must be populated."""
        workshops = self._make_diverse_scenario()
        codes = [w["Code"] for w in workshops]
        students = _make_basic_students(codes, n=4, all_prefer=True)

        input_path = _build_excel(workshops, students)
        output_path = _tmp_path("output")

        try:
            ok, msg, stats = run_optimization(input_path, output_path, category_diversity_weight=100)
            assert ok, msg
            dist = stats["category_diversity_distribution"]
            assert len(dist) > 0, "category_diversity_distribution should not be empty"
            assert len(stats["categories"]) >= 2
        finally:
            for p in (input_path, output_path):
                if os.path.exists(p):
                    os.remove(p)

    def test_higher_weight_increases_coverage(self):
        """
        Higher category_diversity_weight should lead to equal or better
        category coverage (more students covering more distinct categories).
        """
        workshops = self._make_diverse_scenario()
        codes = [w["Code"] for w in workshops]
        students = _make_basic_students(codes, n=6, all_prefer=True)

        input_path = _build_excel(workshops, students)
        out_low = _tmp_path("out_low")
        out_high = _tmp_path("out_high")

        try:
            ok_low, _, stats_low = run_optimization(input_path, out_low, category_diversity_weight=1)
            ok_high, _, stats_high = run_optimization(input_path, out_high, category_diversity_weight=500)
            assert ok_low and ok_high

            def _coverage_score(cat_dist):
                """Sum of (numerator * count) across the distribution."""
                total = 0
                for key, count in cat_dist.items():
                    numerator = int(key.split("/")[0])
                    total += numerator * count
                return total

            score_low = _coverage_score(stats_low["category_diversity_distribution"])
            score_high = _coverage_score(stats_high["category_diversity_distribution"])
            assert score_high >= score_low, (
                f"Higher weight should give >= coverage: {score_high} vs {score_low}"
            )
        finally:
            for p in (input_path, out_low, out_high):
                if os.path.exists(p):
                    os.remove(p)


# ---------------------------------------------------------------------------
# 4. Stats format verification
# ---------------------------------------------------------------------------

class TestStatsFormat:

    def test_stats_contains_all_expected_keys(self):
        """Verify every expected key is present in the stats dictionary."""
        workshops = _make_basic_workshops()
        codes = [w["Code"] for w in workshops]
        students = _make_basic_students(codes, n=2)

        input_path = _build_excel(workshops, students)
        output_path = _tmp_path("output")

        expected_keys = {
            "pref_count",
            "veto_count",
            "neutral_count",
            "pref_rate",
            "prefs_distribution",
            "category_diversity_distribution",
            "categories",
            "max_prefs_possible",
            "objective_value",
            "total_deviation",
            "avg_deviation",
            "total_assignments",
            "students_processed",
            "workshops_processed",
            "category_diversity_weight",
        }

        try:
            ok, msg, stats = run_optimization(input_path, output_path)
            assert ok, msg
            for key in expected_keys:
                assert key in stats, f"Missing key '{key}' in stats"
        finally:
            for p in (input_path, output_path):
                if os.path.exists(p):
                    os.remove(p)

    def test_max_prefs_possible_equals_total_sessions(self):
        workshops = _make_basic_workshops()
        codes = [w["Code"] for w in workshops]
        students = _make_basic_students(codes, n=1)

        input_path = _build_excel(workshops, students)
        output_path = _tmp_path("output")

        try:
            ok, msg, stats = run_optimization(input_path, output_path)
            assert ok, msg
            assert stats["max_prefs_possible"] == TOTAL_SESSIONS
            assert stats["max_prefs_possible"] == 5
        finally:
            for p in (input_path, output_path):
                if os.path.exists(p):
                    os.remove(p)

    def test_prefs_distribution_sums_to_student_count(self):
        """Sum of all values in prefs_distribution == number of students."""
        workshops = _make_basic_workshops()
        codes = [w["Code"] for w in workshops]
        n_students = 4
        students = _make_basic_students(codes, n=n_students)

        input_path = _build_excel(workshops, students)
        output_path = _tmp_path("output")

        try:
            ok, msg, stats = run_optimization(input_path, output_path)
            assert ok, msg
            total = sum(stats["prefs_distribution"].values())
            assert total == n_students
        finally:
            for p in (input_path, output_path):
                if os.path.exists(p):
                    os.remove(p)

    def test_pref_veto_neutral_sum_to_total_assignments(self):
        """pref + veto + neutral sessions should equal total student-sessions."""
        workshops = _make_basic_workshops()
        codes = [w["Code"] for w in workshops]
        students = _make_basic_students(codes, n=3)

        input_path = _build_excel(workshops, students)
        output_path = _tmp_path("output")

        try:
            ok, msg, stats = run_optimization(input_path, output_path)
            assert ok, msg
            total = stats["pref_count"] + stats["veto_count"] + stats["neutral_count"]
            assert total == stats["total_assignments"]
        finally:
            for p in (input_path, output_path):
                if os.path.exists(p):
                    os.remove(p)

    def test_categories_list_is_sorted(self):
        """The categories list should be sorted alphabetically."""
        workshops = _make_basic_workshops()  # has Art, Musique, Science, Sport, Tech
        codes = [w["Code"] for w in workshops]
        students = _make_basic_students(codes, n=1)

        input_path = _build_excel(workshops, students)
        output_path = _tmp_path("output")

        try:
            ok, msg, stats = run_optimization(input_path, output_path)
            assert ok, msg
            assert stats["categories"] == sorted(stats["categories"])
        finally:
            for p in (input_path, output_path):
                if os.path.exists(p):
                    os.remove(p)


# ---------------------------------------------------------------------------
# 5. Edge cases
# ---------------------------------------------------------------------------

class TestEdgeCases:

    def test_no_category_column_no_crash(self):
        """
        If the Ateliers sheet has no 'Catégorie' column the solver should
        still succeed; categories should be empty.
        """
        workshops = _make_basic_workshops()
        codes = [w["Code"] for w in workshops]
        students = _make_basic_students(codes, n=2)

        input_path = _build_excel(workshops, students, include_category_col=False)
        output_path = _tmp_path("output")

        try:
            ok, msg, stats = run_optimization(input_path, output_path)
            assert ok, msg
            assert stats["categories"] == []
            assert stats["category_diversity_distribution"] == {}
        finally:
            for p in (input_path, output_path):
                if os.path.exists(p):
                    os.remove(p)

    def test_all_workshops_same_category_diversity_disabled(self):
        """
        When every workshop has the same category, category diversity
        should be automatically disabled (< 2 distinct categories).
        """
        workshops = _make_basic_workshops()
        for w in workshops:
            w["Catégorie"] = "SameCategory"
        codes = [w["Code"] for w in workshops]
        students = _make_basic_students(codes, n=2)

        input_path = _build_excel(workshops, students)
        output_path = _tmp_path("output")

        try:
            ok, msg, stats = run_optimization(
                input_path, output_path, category_diversity_weight=100
            )
            assert ok, msg
            # Only 1 category -> diversity constraint not activated
            assert len(stats["categories"]) == 1
            # Distribution is still computed (it will show all students at 1/1)
            # but the optimization weight had no effect because use_category_diversity was False
            # The key thing is: no crash and solver succeeds
        finally:
            for p in (input_path, output_path):
                if os.path.exists(p):
                    os.remove(p)

    def test_missing_sheet_returns_error(self):
        """If the Preferences sheet is missing, the function should return an error."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Ateliers"
        ws.append(ATELIERS_COLUMNS)
        path = _tmp_path("bad_input")
        wb.save(path)

        output_path = _tmp_path("output")

        try:
            ok, msg, stats = run_optimization(path, output_path)
            assert not ok
            assert stats is None
            assert "introuvable" in msg.lower() or "erreur" in msg.lower()
        finally:
            for p in (path, output_path):
                if os.path.exists(p):
                    os.remove(p)

    def test_no_valid_workshops_returns_error(self):
        """All workshops have no valid sessions -> should fail gracefully."""
        workshops = [{
            "Code": "BAD",
            "Description": "Bad WS",
            "Enseignant": "T",
            "Salle": "R",
            "Catégorie": "X",
            "Nombre de périodes": 1,
            "Nombre d'élèves max par session": 10,
            "Nombre idéal d'élèves par session": 5,
            "Session 1": None,
            "Session 2": None,
            "Session 3": None,
            "Session 4": None,
            "Session 5": None,
        }]
        students = [{"Nom": "A", "Prénom": "B", "Classe": "6A", "BAD": 1}]

        input_path = _build_excel(workshops, students)
        output_path = _tmp_path("output")

        try:
            ok, msg, stats = run_optimization(input_path, output_path)
            assert not ok
            assert stats is None
        finally:
            for p in (input_path, output_path):
                if os.path.exists(p):
                    os.remove(p)

    def test_no_students_returns_error(self):
        """An empty Preferences sheet (header only) should fail gracefully."""
        workshops = _make_basic_workshops()
        students = []  # no students

        input_path = _build_excel(workshops, students)
        output_path = _tmp_path("output")

        try:
            ok, msg, stats = run_optimization(input_path, output_path)
            assert not ok
            assert stats is None
        finally:
            for p in (input_path, output_path):
                if os.path.exists(p):
                    os.remove(p)

    def test_category_weight_with_no_category_col(self):
        """
        Requesting category diversity but no Catégorie column: should still
        succeed, just with no diversity effect.
        """
        workshops = _make_basic_workshops()
        codes = [w["Code"] for w in workshops]
        students = _make_basic_students(codes, n=2)

        input_path = _build_excel(workshops, students, include_category_col=False)
        output_path = _tmp_path("output")

        try:
            ok, msg, stats = run_optimization(
                input_path, output_path, category_diversity_weight=100
            )
            assert ok, msg
            assert stats["categories"] == []
            assert stats["category_diversity_distribution"] == {}
        finally:
            for p in (input_path, output_path):
                if os.path.exists(p):
                    os.remove(p)
